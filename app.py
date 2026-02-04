from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepInFrame
from reportlab.lib.units import inch
from datetime import datetime
import io
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import RadarChart, Reference

app = Flask(__name__)
app.secret_key = 'vibecoding-secret-key-2026'

STUDENTS_DB = []
SUBJECTS = ['Computer Science', 'Software Engineering', 'Data Science', 'AI & Machine Learning', 'Web Development', 'Mobile Development']

COMMENT_TEMPLATES = {
    "Introduction": [
        {"score": 5, "text": "Excellent introduction - very engaging and clear"},
        {"score": 4, "text": "Good introduction with clear context"},
        {"score": 3, "text": "Adequate introduction, could be more engaging"},
        {"score": 2, "text": "Basic introduction, lacks detail"},
        {"score": 0, "text": "Weak or missing introduction"}
    ],
    "Project Overview & Objectives": [
        {"score": 5, "text": "Objectives are specific, measurable and well-defined"},
        {"score": 4, "text": "Clear objectives with good measurability"},
        {"score": 3, "text": "Objectives present but somewhat unclear"},
        {"score": 2, "text": "Vague objectives, lacking detail"},
        {"score": 0, "text": "Missing or unclear objectives"}
    ],
    "System Requirements": [
        {"score": 5, "text": "Comprehensive functional and non-functional requirements"},
        {"score": 4, "text": "Good requirements documentation"},
        {"score": 3, "text": "Basic requirements, some details missing"},
        {"score": 2, "text": "Incomplete requirements"},
        {"score": 0, "text": "Missing or inadequate requirements"}
    ],
    "Core Functionality & Feature Implementation": [
        {"score": 10, "text": "Excellent feature implementation with relevant technologies"},
        {"score": 8, "text": "Very good features, well implemented"},
        {"score": 6, "text": "Good functionality, minor issues"},
        {"score": 4, "text": "Adequate features, some problems"},
        {"score": 0, "text": "Limited or incomplete functionality"}
    ],
    "System Design": [
        {"score": 5, "text": "Clear architecture with well-justified technology choices"},
        {"score": 4, "text": "Good design documentation and architecture"},
        {"score": 3, "text": "Adequate design, lacks some depth"},
        {"score": 2, "text": "Basic design, incomplete documentation"},
        {"score": 0, "text": "Poor or missing design"}
    ],
    "Results, Challenges & Discussion": [
        {"score": 5, "text": "Thorough analysis with identified challenges and lessons learned"},
        {"score": 4, "text": "Good discussion of results and challenges"},
        {"score": 3, "text": "Adequate discussion, could be more thorough"},
        {"score": 2, "text": "Basic discussion, limited analysis"},
        {"score": 0, "text": "Minimal or no discussion"}
    ],
    "Outlook (Conclusion, Limitation, Future Work)": [
        {"score": 5, "text": "Comprehensive conclusion with clear future directions"},
        {"score": 4, "text": "Good conclusion with future work ideas"},
        {"score": 3, "text": "Basic conclusion, limited future perspectives"},
        {"score": 2, "text": "Weak conclusion"},
        {"score": 0, "text": "Missing conclusion"}
    ],
    "Feature Completeness + Working": [
        {"score": 10, "text": "All features fully implemented and functional"},
        {"score": 8, "text": "Most features working well"},
        {"score": 6, "text": "Good feature completion, minor issues"},
        {"score": 4, "text": "Some features incomplete"},
        {"score": 0, "text": "Many features missing or non-functional"}
    ],
    "User Interface Design + Usability": [
        {"score": 8, "text": "Excellent UI/UX with intuitive navigation and responsive design"},
        {"score": 6, "text": "Good UI design, mostly usable"},
        {"score": 4, "text": "Adequate UI, some usability issues"},
        {"score": 2, "text": "Poor UI/UX design"},
        {"score": 0, "text": "Very poor or missing UI"}
    ],
    "Code Quality & Documentation": [
        {"score": 4, "text": "Clean, well-documented, modular code"},
        {"score": 3, "text": "Good code quality with adequate documentation"},
        {"score": 2, "text": "Acceptable code, limited documentation"},
        {"score": 1, "text": "Poor code quality or minimal documentation"},
        {"score": 0, "text": "Very poor code quality or no documentation"}
    ],
    "Containerized DevOps": [
        {"score": 6, "text": "Excellent Docker setup, runs smoothly with docker-compose"},
        {"score": 5, "text": "Good containerization with minor issues"},
        {"score": 3, "text": "Basic containerization, some problems"},
        {"score": 1, "text": "Poor containerization"},
        {"score": 0, "text": "Non-functional or missing containerization"}
    ],
    "Testing & Reliability": [
        {"score": 2, "text": "Comprehensive tests, robust error handling"},
        {"score": 1, "text": "Basic testing, some error handling"},
        {"score": 0, "text": "No testing or error handling"}
    ],
    "Presentation": [
        {"score": 15, "text": "Exceptional delivery, highly engaging and professional"},
        {"score": 12, "text": "Confident delivery, minor pace or emphasis issues"},
        {"score": 9, "text": "Clear delivery, lacks polish"},
        {"score": 5, "text": "Needs improvement in clarity and delivery"},
        {"score": 0, "text": "Weak or missing presentation"}
    ],
    "Peer reviews & participation": [
        {"score": 10, "text": "The student actively participated in peer reviews, forums, and live sessions, providing meaningful and constructive feedback."},
        {"score": 9, "text": "The student participated regularly but could have contributed more depth or frequency to their feedback."},
        {"score": 7, "text": "The student participated occasionally, with limited or superficial contributions."},
        {"score": 5, "text": "The student participated minimally, with little to no meaningful feedback."},
        {"score": 0, "text": "No meaningful participation in peer activities."}
    ]
}

# Report-type specific overrides for comment templates
REPORT_SPECIFIC_TEMPLATES = {
    "Seminar Report": {
        "Core Functionality & Feature Implementation": [
            {"score": 10, "text": "Outstanding live presentation and demonstration of features"},
            {"score": 8, "text": "Strong demonstration during seminar with most features present"},
            {"score": 6, "text": "Adequate demonstration, some features missing"},
            {"score": 4, "text": "Limited demonstration; important parts missing"},
            {"score": 0, "text": "No demonstration during seminar"}
        ],
        "Results, Challenges & Discussion": [
            {"score": 5, "text": "Excellent discussion with clear reflections on the seminar work"},
            {"score": 4, "text": "Good discussion, some depth missing"},
            {"score": 3, "text": "Basic discussion; limited reflection"},
            {"score": 1, "text": "Minimal discussion"},
            {"score": 0, "text": "No discussion presented"}
        ],
        "Presentation": [
            {"score": 15, "text": "Exceptional delivery, highly engaging and professional"},
            {"score": 12, "text": "Confident delivery, minor pace or emphasis issues"},
            {"score": 9, "text": "Clear delivery, lacks polish"},
            {"score": 5, "text": "Needs improvement in clarity and delivery"},
            {"score": 0, "text": "Weak or missing presentation"}
        ],
        "Report": [
            {"score": 75, "text": "Excellent report quality and completeness"},
            {"score": 60, "text": "Good report with minor issues"},
            {"score": 45, "text": "Adequate report, several gaps"},
            {"score": 25, "text": "Weak report"},
            {"score": 0, "text": "No meaningful report"}
        ],
        "Peer Activities": [
            {"score": 10, "text": "Strong participation and constructive feedback"},
            {"score": 8, "text": "Good participation with minor gaps"},
            {"score": 6, "text": "Limited participation"},
            {"score": 4, "text": "Minimal participation"},
            {"score": 0, "text": "No participation"}
        ]
    },
    "Research-driven Thesis": {
        "Results, Challenges & Discussion": [
            {"score": 5, "text": "In-depth analysis and methodological rigor"},
            {"score": 4, "text": "Solid analysis with adequate methodology"},
            {"score": 3, "text": "Some analysis but lacks depth"},
            {"score": 1, "text": "Shallow analysis"},
            {"score": 0, "text": "No meaningful analysis"}
        ],
        "System Design": [
            {"score": 5, "text": "Clear theoretical framework and design choices"},
            {"score": 4, "text": "Good design rationale"},
            {"score": 3, "text": "Functional design, needs justification"},
            {"score": 1, "text": "Weak design reasoning"},
            {"score": 0, "text": "No design rationale"}
        ],
        "General Aspects": [
            {"score": 30, "text": "Excellent structure, clarity, and overall quality"},
            {"score": 24, "text": "Good overall quality"},
            {"score": 18, "text": "Adequate, needs improvement"},
            {"score": 10, "text": "Weak overall quality"},
            {"score": 0, "text": "Very poor overall quality"}
        ],
        "Theory & Framework, Literature Review or Research Questions": [
            {"score": 30, "text": "Strong theory and research framing"},
            {"score": 24, "text": "Good theoretical foundation"},
            {"score": 18, "text": "Basic theory, lacks depth"},
            {"score": 10, "text": "Weak theoretical basis"},
            {"score": 0, "text": "Missing theoretical basis"}
        ],
        "Research": [
            {"score": 40, "text": "Excellent research design and execution"},
            {"score": 32, "text": "Good research with minor issues"},
            {"score": 24, "text": "Adequate research, limited depth"},
            {"score": 12, "text": "Weak research"},
            {"score": 0, "text": "No meaningful research"}
        ]
    },
    "Design-driven + Small Evaluation": {
        "System Development": [
            {"score": 70, "text": "Excellent system development with strong execution"},
            {"score": 55, "text": "Good system development with minor issues"},
            {"score": 40, "text": "Adequate development, limited depth"},
            {"score": 20, "text": "Weak or incomplete system development"},
            {"score": 0, "text": "No meaningful system development"}
        ],
        "Evaluation": [
            {"score": 20, "text": "Excellent evaluation methodology and results"},
            {"score": 16, "text": "Good evaluation with minor issues"},
            {"score": 12, "text": "Basic evaluation, limited depth"},
            {"score": 6, "text": "Weak evaluation"},
            {"score": 0, "text": "No evaluation provided"}
        ],
        "General Aspects": [
            {"score": 30, "text": "Excellent structure, clarity, and overall quality"},
            {"score": 24, "text": "Good overall quality"},
            {"score": 18, "text": "Adequate, needs improvement"},
            {"score": 10, "text": "Weak overall quality"},
            {"score": 0, "text": "Very poor overall quality"}
        ]
    }
}

# Provide basic templates for new rubric categories if not present
COMMENT_TEMPLATES.setdefault("General Aspects", [
    {"score": 30, "text": "Excellent structure, clarity, and overall quality"},
    {"score": 24, "text": "Good overall quality with minor issues"},
    {"score": 18, "text": "Adequate quality, needs improvement"},
    {"score": 10, "text": "Weak overall quality"},
    {"score": 0, "text": "Very poor overall quality"}
])
COMMENT_TEMPLATES.setdefault("Theory & Framework, Literature Review or Research Questions", [
    {"score": 30, "text": "Strong theory and research framing"},
    {"score": 24, "text": "Good theoretical foundation"},
    {"score": 18, "text": "Basic theory, lacks depth"},
    {"score": 10, "text": "Weak theoretical basis"},
    {"score": 0, "text": "Missing theoretical basis"}
])
COMMENT_TEMPLATES.setdefault("Research", [
    {"score": 40, "text": "Excellent research design and execution"},
    {"score": 32, "text": "Good research with minor issues"},
    {"score": 24, "text": "Adequate research, limited depth"},
    {"score": 12, "text": "Weak research"},
    {"score": 0, "text": "No meaningful research"}
])
COMMENT_TEMPLATES.setdefault("System Development", [
    {"score": 70, "text": "Excellent system development with strong execution"},
    {"score": 55, "text": "Good system development with minor issues"},
    {"score": 40, "text": "Adequate development, limited depth"},
    {"score": 20, "text": "Weak system development"},
    {"score": 0, "text": "No meaningful system development"}
])
COMMENT_TEMPLATES.setdefault("Evaluation", [
    {"score": 20, "text": "Excellent evaluation methodology and results"},
    {"score": 16, "text": "Good evaluation with minor issues"},
    {"score": 12, "text": "Basic evaluation, limited depth"},
    {"score": 6, "text": "Weak evaluation"},
    {"score": 0, "text": "No evaluation provided"}
])
COMMENT_TEMPLATES.setdefault("Modeling", [
    {"score": 50, "text": "Excellent modeling with strong justification"},
    {"score": 40, "text": "Good modeling with minor issues"},
    {"score": 30, "text": "Adequate modeling, limited depth"},
    {"score": 15, "text": "Weak modeling"},
    {"score": 0, "text": "No meaningful modeling"}
])
COMMENT_TEMPLATES.setdefault("Report", [
    {"score": 75, "text": "Excellent report quality and completeness"},
    {"score": 60, "text": "Good report with minor issues"},
    {"score": 45, "text": "Adequate report, several gaps"},
    {"score": 25, "text": "Weak report"},
    {"score": 0, "text": "No meaningful report"}
])
COMMENT_TEMPLATES.setdefault("Peer Activities", [
    {"score": 10, "text": "Strong participation and constructive feedback"},
    {"score": 8, "text": "Good participation with minor gaps"},
    {"score": 6, "text": "Limited participation"},
    {"score": 4, "text": "Minimal participation"},
    {"score": 0, "text": "No participation"}
])

def get_comment_templates_for_report(report_type):
    templates = COMMENT_TEMPLATES.copy()
    overrides = REPORT_SPECIFIC_TEMPLATES.get(report_type, {})
    for cat, items in overrides.items():
        templates[cat] = items
    return templates

# Presentation-specific rubric
PRESENTATION_RUBRIC = {
    "Presentation": {
        "max_points": 15,
        "criteria": [
            {"range": "14-15", "description": "Exceptional delivery, highly engaging and professional"},
            {"range": "12-13", "description": "Confident delivery, minor pacing or emphasis issues"},
            {"range": "10-11", "description": "Clear delivery but lacks polish"},
            {"range": "0-9", "description": "Delivery needs improvement"}
        ]
    }
}

# Base groups (used for grouping titles in the UI)
GROUPS = {
    "Presentation": ["Presentation"],
    "Report": [
        "General Aspects",
        "Theory & Framework, Literature Review or Research Questions",
        "Research",
        "System Development",
        "Evaluation",
        "Modeling",
        "Report"
    ],
    "Peer activities": [
        "Peer Activities"
    ]
}

# Report-type to exact categories mapping
REPORT_CATEGORY_MAP = {
    "Research-driven Thesis": [
        "General Aspects",
        "Theory & Framework, Literature Review or Research Questions",
        "Research"
    ],
    "Design-driven + Small Evaluation": [
        "General Aspects",
        "System Development",
        "Evaluation"
    ],
    "Design-driven Thesis": [
        "General Aspects",
        "System Development"
    ],
    "ML/NLP Thesis": [
        "General Aspects",
        "Modeling",
        "Evaluation"
    ],
    "Seminar Report": [
        "Presentation",
        "Report",
        "Peer Activities"
    ]
}

# Rubrics with report-type specific categories
RUBRICS = {
    "Presentation": {
        "max_points": 15,
        "criteria": [
            {"range": "14-15", "description": "Exceptional delivery, highly engaging and professional"},
            {"range": "12-13", "description": "Confident delivery, minor pacing or emphasis issues"},
            {"range": "10-11", "description": "Clear delivery but lacks polish"},
            {"range": "0-9", "description": "Delivery needs improvement"}
        ]
    },
    "General Aspects": {
        "max_points": 30,
        "criteria": [
            {"range": "27-30", "description": "Excellent structure, clarity, and overall quality"},
            {"range": "21-26", "description": "Good quality with minor issues"},
            {"range": "15-20", "description": "Adequate quality, several gaps"},
            {"range": "0-14", "description": "Weak or incomplete overall quality"}
        ]
    },
    "Theory & Framework, Literature Review or Research Questions": {
        "max_points": 30,
        "criteria": [
            {"range": "27-30", "description": "Strong theory and research framing"},
            {"range": "21-26", "description": "Good theoretical foundation"},
            {"range": "15-20", "description": "Basic theory, lacks depth"},
            {"range": "0-14", "description": "Weak or missing theoretical basis"}
        ]
    },
    "Research": {
        "max_points": 40,
        "criteria": [
            {"range": "36-40", "description": "Excellent research design and execution"},
            {"range": "30-35", "description": "Good research with minor issues"},
            {"range": "20-29", "description": "Adequate research, limited depth"},
            {"range": "0-19", "description": "Weak or missing research"}
        ]
    },
    "System Development": {
        "max_points": 70,
        "criteria": [
            {"range": "63-70", "description": "Excellent system design and implementation"},
            {"range": "50-62", "description": "Good implementation with minor issues"},
            {"range": "35-49", "description": "Adequate implementation, several gaps"},
            {"range": "0-34", "description": "Weak or incomplete system development"}
        ]
    },
    "Evaluation": {
        "max_points": 20,
        "criteria": [
            {"range": "18-20", "description": "Excellent evaluation methodology and results"},
            {"range": "14-17", "description": "Good evaluation with minor issues"},
            {"range": "10-13", "description": "Basic evaluation, limited depth"},
            {"range": "0-9", "description": "Weak or missing evaluation"}
        ]
    },
    "Modeling": {
        "max_points": 50,
        "criteria": [
            {"range": "45-50", "description": "Excellent modeling with strong justification"},
            {"range": "38-44", "description": "Good modeling with minor issues"},
            {"range": "25-37", "description": "Adequate modeling, limited depth"},
            {"range": "0-24", "description": "Weak or missing modeling"}
        ]
    },
    "Report": {
        "max_points": 75,
        "criteria": [
            {"range": "68-75", "description": "Excellent report quality and completeness"},
            {"range": "55-67", "description": "Good report with minor issues"},
            {"range": "40-54", "description": "Adequate report, several gaps"},
            {"range": "0-39", "description": "Weak or incomplete report"}
        ]
    },
    "Peer Activities": {
        "max_points": 10,
        "criteria": [
            {"range": "9-10", "description": "Strong participation and constructive feedback"},
            {"range": "7-8", "description": "Good participation with minor gaps"},
            {"range": "5-6", "description": "Limited participation"},
            {"range": "0-4", "description": "Minimal or no participation"}
        ]
    }
}

@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    username = request.form.get('username')
    password = request.form.get('password')
    if username and password:
        session['user'] = username
        return redirect(url_for('dashboard'))
    return redirect(url_for('index'))

@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect(url_for('index'))
    return render_template('dashboard.html', user=session['user'])

@app.route('/evaluate')
def evaluate():
    if 'user' not in session:
        return redirect(url_for('index'))
    # Step 1: student details page
    return render_template('evaluate_details.html', subjects=SUBJECTS, current=session.get('current_evaluation', {}))

@app.route('/evaluate/start', methods=['POST'])
def evaluate_start():
    if 'user' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json() or {}
    session['current_evaluation'] = {
        'first_name': data.get('first_name', ''),
        'last_name': data.get('last_name', ''),
        'student_id': data.get('student_id', ''),
        'gender': data.get('gender', ''),
        'report': data.get('report', ''),
        'report_type': data.get('report_type', ''),
        'report_title': data.get('report_title', ''),
        'seminar_date': data.get('seminar_date', ''),
        'seminar_time': data.get('seminar_time', '')
    }
    return jsonify({'status': 'ok'})

@app.route('/evaluate/grade')
def evaluate_grade():
    if 'user' not in session:
        return redirect(url_for('index'))
    current = session.get('current_evaluation')
    if not current:
        return redirect(url_for('evaluate'))
    # Provide comment templates tailored to the chosen report type
    ct = get_comment_templates_for_report(current.get('report_type', ''))
    rubrics, groups = get_rubrics_and_groups(current.get('report_type', ''))
    return render_template('evaluate_grade.html', rubrics=rubrics, comment_templates=ct, current=current, groups=groups)

@app.route('/evaluate/clear_session', methods=['POST'])
def evaluate_clear_session():
    session.pop('current_evaluation', None)
    return jsonify({'status': 'cleared'})

@app.route('/evaluate/current')
def evaluate_current():
    """Return the current evaluation stored in session as JSON."""
    if 'user' not in session:
        return jsonify({}), 401
    return jsonify(session.get('current_evaluation', {}))

@app.route('/students')
def students():
    if 'user' not in session:
        return redirect(url_for('index'))
    return render_template('students.html', students=STUDENTS_DB)

@app.route('/statistics')
def statistics():
    if 'user' not in session:
        return redirect(url_for('index'))
    return render_template('statistics.html', students=STUDENTS_DB)

@app.route('/settings')
def settings():
    if 'user' not in session:
        return redirect(url_for('index'))
    return render_template('settings.html')

@app.route('/admin')
def admin():
    if 'user' not in session:
        return redirect(url_for('index'))
    return render_template('admin.html', rubrics=RUBRICS)

@app.route('/admin/rubric/update', methods=['POST'])
def admin_update_rubric():
    if 'user' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json() or {}
    category = (data.get('category') or '').strip()
    max_points = data.get('max_points')
    criteria = data.get('criteria') or []
    if not category:
        return jsonify({'error': 'Missing category'}), 400
    try:
        max_points = int(max_points)
    except Exception:
        return jsonify({'error': 'Invalid max_points'}), 400
    norm_criteria = []
    for c in criteria:
        if not isinstance(c, dict):
            continue
        rng = str(c.get('range', '')).strip()
        desc = str(c.get('description', '')).strip()
        if rng and desc:
            norm_criteria.append({'range': rng, 'description': desc})
    RUBRICS[category] = {'max_points': max_points, 'criteria': norm_criteria}
    # ensure category is in a group (default to Report)
    if category not in sum(GROUPS.values(), []):
        GROUPS.setdefault('Report', []).append(category)
    return jsonify({'status': 'ok'})

@app.route('/admin/rubric/delete', methods=['POST'])
def admin_delete_rubric():
    if 'user' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json() or {}
    category = (data.get('category') or '').strip()
    if not category or category not in RUBRICS:
        return jsonify({'error': 'Not found'}), 404
    RUBRICS.pop(category, None)
    for group, cats in list(GROUPS.items()):
        if category in cats:
            GROUPS[group] = [c for c in cats if c != category]
        if not GROUPS[group]:
            GROUPS.pop(group, None)
    return jsonify({'status': 'deleted'})

@app.route('/export_csv')
def export_csv():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['First Name', 'Last Name', 'ID', 'Gender', 'Report Type', 'Report Title', 'Seminar Date', 'Seminar Time', 'Total Score', 'Percentage', 'Grade', 'Date'])
    
    for student in STUDENTS_DB:
        writer.writerow([
            student.get('first_name', 'N/A'),
            student.get('last_name', 'N/A'),
            student['id'],
            student.get('gender', 'N/A'),
            student.get('report_type', 'N/A'),
            student.get('report_title', 'N/A'),
            student.get('seminar_date', 'N/A'),
            student.get('seminar_time', 'N/A'),
            student['total_score'],
            f"{student['percentage']:.1f}%",
            student['grade'],
            student['date']
        ])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'evaluations_{datetime.now().strftime("%Y%m%d")}.csv'
    )

@app.route('/export_excel')
def export_excel():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Evaluations'
    
    headers = ['First Name', 'Last Name', 'ID', 'Gender', 'Report Type', 'Total Score', 'Percentage', 'Grade', 'Date']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for student in STUDENTS_DB:
        ws.append([
            student.get('first_name', 'N/A'),
            student.get('last_name', 'N/A'),
            student['id'],
            student.get('gender', 'N/A'),
            student.get('report_type', 'N/A'),
            student['total_score'],
            f"{student['percentage']:.1f}%",
            student['grade'],
            student['date']
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'evaluations_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/compare')
def compare():
    if 'user' not in session:
        return redirect(url_for('index'))
    return render_template('compare.html', students=STUDENTS_DB, rubrics=RUBRICS)

@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('index'))

@app.route('/generate_report', methods=['POST'])
def generate_report():
    data = request.json
    rubrics, groups = get_rubrics_and_groups(data.get('report_type', ''))
    record = _build_student_record(data, rubrics=rubrics, groups=groups)
    _upsert_student(record)

    student_name = record.get('name', 'N/A')
    student_id = record.get('id', 'N/A')
    gender = record.get('gender', 'N/A')
    report_type = record.get('report_type', 'N/A')
    report_title = record.get('report_title', 'N/A')
    seminar_date = record.get('seminar_date', 'N/A')
    seminar_time = record.get('seminar_time', 'N/A')
    scores = record.get('scores_raw', {})
    comments = record.get('comments', {})
    total_score = record.get('total_score', 0)
    max_total = sum(meta.get('max_points', 0) for meta in rubrics.values()) or 1
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=30,
        alignment=1
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    story.append(Paragraph("📊 Project Evaluation Report", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    info_data = [
        ['Student Name:', student_name],
        ['Student ID:', student_id],
        ['Gender:', gender],
        ['Report Type:', report_type],
        ['Report Title:', report_title],
        ['Seminar Date:', seminar_date if seminar_date != 'N/A' else 'Not specified'],
        ['Seminar Time:', seminar_time if seminar_time != 'N/A' else 'Not specified'],
        ['Evaluation Date:', datetime.now().strftime('%d.%m.%Y')],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    total_score = 0
    max_total = sum(meta.get('max_points', 0) for meta in RUBRICS.values()) or 1
    
    story.append(Paragraph("Detailed Evaluation", heading_style))
    story.append(Spacer(1, 0.1*inch))

    comment_style = ParagraphStyle(
        'CommentStyle',
        parent=styles['BodyText'],
        fontSize=9,
        leading=11,
        textColor=colors.black
    )

    def normalize_comment(value):
        if value is None:
            return ''
        if isinstance(value, list):
            if value and isinstance(value[0], dict):
                return value[0].get('comment') or value[0].get('text') or ''
            return ', '.join(str(c) for c in value if c is not None)
        if isinstance(value, dict):
            return value.get('comment') or value.get('text') or ''
        return str(value)

    # Render grouped sections (Presentation, Report, ...)
    for group, cats in groups.items():
        story.append(Paragraph(group, styles['Heading3']))
        story.append(Spacer(1, 0.05*inch))
        group_total = 0
        group_max = 0
        for category in cats:
            details = rubrics.get(category)
            if not details:
                continue
            raw_score = scores.get(category, 0)
            score = _extract_points(raw_score)
            group_total += score
            group_max += details.get('max_points', 0)

            comment = normalize_comment(comments.get(category, ''))
            comment_para = Paragraph(comment, comment_style) if comment else Paragraph('', comment_style)
            comment_cell = KeepInFrame(4.5*inch, 0.9*inch, [comment_para], mode='shrink')

            story.append(Paragraph(f"<b>{category}</b>", styles['Heading4']))
            score_data = [
                ['Score:', f"{score} / {details.get('max_points',0)}"],
                ['Comment:', comment_cell]
            ]
            score_table = Table(score_data, colWidths=[1.5*inch, 4.5*inch])
            score_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP')
            ]))
            story.append(score_table)
            story.append(Spacer(1, 0.08*inch))

        story.append(Paragraph(f"<b>{group} Subtotal: {group_total} / {group_max}</b>", styles['BodyText']))
        story.append(Spacer(1, 0.12*inch))
        total_score += group_total
    
    for category, details in rubrics.items():
        # legacy loop disabled (grouped rendering above)
        continue
        raw_score = scores.get(category, 0)
        score = _extract_points(raw_score)
        total_score += score
        comment = normalize_comment(comments.get(category, ''))
        comment_para = Paragraph(comment, comment_style) if comment else Paragraph('', comment_style)
        comment_cell = KeepInFrame(4.5*inch, 0.9*inch, [comment_para], mode='shrink')
        
        story.append(Paragraph(f"<b>{category}</b>", styles['Heading3']))
        
        score_data = [
            ['Score:', f"{score} / {details['max_points']}"],
            ['Comment:', comment_cell]
        ]
        
        score_table = Table(score_data, colWidths=[1.5*inch, 4.5*inch])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP')
        ]))
        
        story.append(score_table)
        story.append(Spacer(1, 0.15*inch))
    
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph("Final Summary", heading_style))
    
    percentage = (total_score / max_total) * 100
    grade = get_grade(percentage)
    
    summary_data = [
        ['Total Score:', f"{total_score} / {max_total}"],
        ['Percentage:', f"{percentage:.1f}%"],
        ['Grade:', grade]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dbeafe')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2563eb'))
    ]))
    
    story.append(summary_table)
    
    doc.build(story)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name=f'evaluation_{student_id}.pdf', mimetype='application/pdf')

@app.route('/evaluate/save', methods=['POST'])
def evaluate_save():
    if 'user' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json() or {}
    rubrics, groups = get_rubrics_and_groups(data.get('report_type', ''))
    record = _build_student_record(data, rubrics=rubrics, groups=groups)
    index, created = _upsert_student(record)
    return jsonify({'status': 'ok', 'index': index, 'created': created})

@app.route('/students/delete/<int:idx>', methods=['POST'])
def delete_student(idx):
    if 'user' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    if idx < 0 or idx >= len(STUDENTS_DB):
        return jsonify({'error': 'Not found'}), 404
    STUDENTS_DB.pop(idx)
    return jsonify({'status': 'deleted'})

def get_grade(percentage):
    if percentage >= 90:
        return "Excellent (1.0-1.3)"
    elif percentage >= 80:
        return "Very Good (1.7-2.3)"
    elif percentage >= 70:
        return "Good (2.7-3.3)"
    elif percentage >= 60:
        return "Satisfactory (3.7-4.0)"
    else:
        return "Fail (5.0)"

def get_rubrics_and_groups(report_type):
    categories = REPORT_CATEGORY_MAP.get(report_type)
    if not categories:
        categories = [c for cats in GROUPS.values() for c in cats]
    # build grouped view
    if report_type == "Seminar Report":
        groups = {
            "Presentation": [c for c in categories if c == "Presentation"],
            "Report": [c for c in categories if c == "Report"],
            "Peer activities": [c for c in categories if c == "Peer Activities"]
        }
    else:
        groups = {"Report": [c for c in categories if c in RUBRICS]}
    groups = {g: cats for g, cats in groups.items() if cats}
    allowed = {c for cats in groups.values() for c in cats}
    rubrics = {c: RUBRICS[c] for c in allowed if c in RUBRICS}
    return rubrics, groups

def _extract_points(value):
    if isinstance(value, list):
        total = 0
        for item in value:
            if isinstance(item, dict):
                total += float(item.get('points', 0) or 0)
            else:
                try:
                    total += float(item)
                except Exception:
                    pass
        return total
    try:
        return float(value)
    except Exception:
        return 0

def _build_student_record(data, rubrics=None, groups=None):
    rubrics = rubrics or RUBRICS
    groups = groups or GROUPS
    first_name = data.get('first_name', 'N/A')
    last_name = data.get('last_name', 'N/A')
    student_name = f"{first_name} {last_name}"
    student_id = data.get('student_id', 'N/A')
    gender = data.get('gender', 'N/A')
    report_type = data.get('report_type', 'N/A')
    report_title = data.get('report_title', 'N/A')
    seminar_date = data.get('seminar_date', 'N/A')
    seminar_time = data.get('seminar_time', 'N/A')
    report = data.get('report', 'N/A')
    scores = data.get('scores', {})
    comments = data.get('comments', {})

    normalized_scores = {}
    for cat in rubrics.keys():
        normalized_scores[cat] = _extract_points(scores.get(cat, 0))

    total_score = sum(normalized_scores.values())
    max_total = sum(meta.get('max_points', 0) for meta in rubrics.values()) or 1
    percentage = (total_score / max_total) * 100
    grade = get_grade(percentage)

    group_scores = {}
    for group, cats in groups.items():
        group_scores[group] = sum(normalized_scores.get(c, 0) for c in cats)

    return {
        'name': student_name,
        'first_name': first_name,
        'last_name': last_name,
        'id': student_id,
        'gender': gender,
        'report': report,
        'report_type': report_type,
        'report_title': report_title,
        'seminar_date': seminar_date,
        'seminar_time': seminar_time,
        'total_score': total_score,
        'percentage': percentage,
        'grade': grade,
        'date': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'scores': normalized_scores,
        'scores_raw': scores,
        'comments': comments,
        'group_scores': group_scores
    }

def _upsert_student(record):
    key = (record.get('id'), record.get('report_type'), record.get('report_title'))
    for i, s in enumerate(STUDENTS_DB):
        if (s.get('id'), s.get('report_type'), s.get('report_title')) == key:
            STUDENTS_DB[i] = record
            return i, False
    STUDENTS_DB.append(record)
    return len(STUDENTS_DB) - 1, True

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
